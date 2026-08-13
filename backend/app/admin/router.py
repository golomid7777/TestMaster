from fastapi import APIRouter, HTTPException, Depends
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from datetime import datetime, timedelta

from database.db import SessionLocal
from database.models import Topic, Service, User, Question, TopicAccess
from security import decode_token

from fastapi import UploadFile, File
from openpyxl import load_workbook
from io import BytesIO

from fastapi.responses import StreamingResponse
from openpyxl import Workbook

router = APIRouter(
    prefix="/admin",
    tags=["Admin"]
)


security = HTTPBearer()


def require_admin(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    payload = decode_token(credentials.credentials)

    if not payload:
        raise HTTPException(
            status_code=401,
            detail="Invalid token"
        )

    if not payload.get("is_admin"):
        raise HTTPException(
            status_code=403,
            detail="Admin access required"
        )

    return payload


class ServiceCreate(BaseModel):
    name: str


class TopicCreate(BaseModel):
    name: str
    service_id: int
    time_limit_minutes: int = 30
    price_kopecks: int = 0
    access_minutes: int = 0

class TopicUpdate(BaseModel):
    name: str
    time_limit_minutes: int
    price_kopecks: int = 0
    access_minutes: int = 0


class UserServiceUpdate(BaseModel):
    service_id: int


class GrantTopicAccess(BaseModel):
    minutes: int


# =========================================================
# СЛУЖБЫ
# =========================================================

@router.get("/services")
def get_services(
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        services = db.query(Service).order_by(
            Service.name
        ).all()

        return [
            {
                "id": service.id,
                "name": service.name
            }
            for service in services
        ]

    finally:
        db.close()


@router.post("/services")
def create_service(
    data: ServiceCreate,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        name = data.name.strip()

        if not name:
            raise HTTPException(
                status_code=400,
                detail="Название службы не указано"
            )

        exists = db.query(Service).filter(
            Service.name == name
        ).first()

        if exists:
            raise HTTPException(
                status_code=400,
                detail="Такая служба уже существует"
            )

        service = Service(
            name=name
        )

        db.add(service)
        db.commit()
        db.refresh(service)

        return {
            "id": service.id,
            "name": service.name
        }

    finally:
        db.close()


@router.delete("/services/{service_id}")
def delete_service(
    service_id: int,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        service = db.query(Service).filter(
            Service.id == service_id
        ).first()

        if not service:
            raise HTTPException(
                status_code=404,
                detail="Служба не найдена"
            )

        user = db.query(User).filter(
            User.service_id == service_id
        ).first()

        if user:
            raise HTTPException(
                status_code=400,
                detail="Нельзя удалить службу: к ней привязаны пользователи"
            )

        if service.questions:
            raise HTTPException(
                status_code=400,
                detail="Нельзя удалить службу: в ней есть вопросы"
            )

        if service.topics:
            raise HTTPException(
                status_code=400,
                detail="Сначала удалите темы этой службы"
            )

        db.delete(service)
        db.commit()

        return {
            "status": "deleted"
        }

    finally:
        db.close()


# =========================================================
# ТЕМЫ
# =========================================================

@router.get("/topics")
def get_topics(
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        topics = (
            db.query(Topic, Service)
            .join(
                Service,
                Topic.service_id == Service.id
            )
            .order_by(
                Service.name,
                Topic.name
            )
            .all()
        )

        return [
            {
                "id": topic.id,
                "name": topic.name,
                "service_id": topic.service_id,
                "service_name": service.name,
                "time_limit_minutes": topic.time_limit_minutes,
                "price_kopecks": topic.price_kopecks,
                "access_minutes": topic.access_minutes
            }
            for topic, service in topics
        ]

    finally:
        db.close()


@router.post("/topics")
def create_topic(
    data: TopicCreate,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        service = db.query(Service).filter(
            Service.id == data.service_id
        ).first()

        if not service:
            raise HTTPException(
                status_code=404,
                detail="Служба не найдена"
            )

        name = data.name.strip()

        if not name:
            raise HTTPException(
                status_code=400,
                detail="Название темы не указано"
            )

        exists = db.query(Topic).filter(
            Topic.name == name,
            Topic.service_id == data.service_id
        ).first()

        if exists:
            raise HTTPException(
                status_code=400,
                detail="Такая тема уже существует"
            )

        if data.time_limit_minutes < 1:
            raise HTTPException(
           status_code=400,
           detail="Время теста должно быть не меньше 1 минуты"
        )

        if data.price_kopecks < 0:
            raise HTTPException(
                status_code=400,
                detail="Цена не может быть отрицательной"
            )

        if data.price_kopecks > 0 and data.access_minutes < 1:
            raise HTTPException(
                status_code=400,
                detail="Для платной темы укажите срок доступа в минутах"
            )

        topic = Topic(
           name=name,
           service_id=data.service_id,
           time_limit_minutes=data.time_limit_minutes,
           price_kopecks=data.price_kopecks,
           access_minutes=data.access_minutes
       )

        db.add(topic)
        db.commit()
        db.refresh(topic)

        return {
            "id": topic.id,
            "name": topic.name,
            "service_id": topic.service_id,
            "time_limit_minutes": topic.time_limit_minutes,
            "price_kopecks": topic.price_kopecks,
            "access_minutes": topic.access_minutes
        }

    finally:
        db.close()

@router.put("/topics/{topic_id}")
def update_topic(
    topic_id: int,
    data: TopicUpdate,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        topic = db.query(Topic).filter(
            Topic.id == topic_id
        ).first()

        if not topic:
            raise HTTPException(
                status_code=404,
                detail="Тема не найдена"
            )

        name = data.name.strip()

        if not name:
            raise HTTPException(
                status_code=400,
                detail="Введите название темы"
            )

        if data.time_limit_minutes < 1:
            raise HTTPException(
                status_code=400,
                detail="Время теста должно быть не меньше 1 минуты"
            )

        if data.price_kopecks < 0:
            raise HTTPException(
                status_code=400,
                detail="Цена не может быть отрицательной"
            )

        if data.price_kopecks > 0 and data.access_minutes < 1:
            raise HTTPException(
                status_code=400,
                detail="Для платной темы укажите срок доступа в минутах"
            )

        topic.name = name
        topic.time_limit_minutes = data.time_limit_minutes
        topic.price_kopecks = data.price_kopecks
        topic.access_minutes = data.access_minutes

        db.commit()
        db.refresh(topic)

        return {
            "id": topic.id,
            "name": topic.name,
            "service_id": topic.service_id,
            "time_limit_minutes": topic.time_limit_minutes,
            "price_kopecks": topic.price_kopecks,
            "access_minutes": topic.access_minutes
        }

    finally:
        db.close()

@router.delete("/topics/{topic_id}")
def delete_topic(
    topic_id: int,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        topic = db.query(Topic).filter(
            Topic.id == topic_id
        ).first()

        if not topic:
            raise HTTPException(
                status_code=404,
                detail="Тема не найдена"
            )

        if topic.questions:
            raise HTTPException(
                status_code=400,
                detail="Нельзя удалить тему: в ней есть вопросы"
            )

        db.delete(topic)
        db.commit()

        return {
            "status": "deleted"
        }

    finally:
        db.close()


# =========================================================
# ПОЛЬЗОВАТЕЛИ
# =========================================================

@router.get("/users")
def get_users(
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        users = (
            db.query(User)
            .order_by(User.id)
            .all()
        )

        result = []

        for user in users:

            service = db.query(Service).filter(
                Service.id == user.service_id
            ).first()

            result.append({
                "id": user.id,
                "name": user.name,
                "email": user.email,
                "service_id": user.service_id,
                "service_name": (
                    service.name
                    if service
                    else "Служба не найдена"
                ),
                "is_admin": bool(user.is_admin)
            })

        return result

    finally:
        db.close()


@router.put("/users/{user_id}/service")
def change_user_service(
    user_id: int,
    data: UserServiceUpdate,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        user = db.query(User).filter(
            User.id == user_id
        ).first()

        if not user:
            raise HTTPException(
                status_code=404,
                detail="Пользователь не найден"
            )

        service = db.query(Service).filter(
            Service.id == data.service_id
        ).first()

        if not service:
            raise HTTPException(
                status_code=404,
                detail="Служба не найдена"
            )

        user.service_id = data.service_id

        db.commit()

        return {
            "status": "updated",
            "user_id": user.id,
            "service_id": user.service_id
        }

    finally:
        db.close()


@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        current_admin_id = admin.get("user_id")

        try:
            current_admin_id = int(current_admin_id)
        except (TypeError, ValueError):
            pass

        if current_admin_id == user_id:
            raise HTTPException(
                status_code=400,
                detail="Нельзя удалить собственный аккаунт администратора"
            )

        user = db.query(User).filter(
            User.id == user_id
        ).first()

        if not user:
            raise HTTPException(
                status_code=404,
                detail="Пользователь не найден"
            )

        db.delete(user)
        db.commit()

        return {
            "status": "deleted"
        }

    finally:
        db.close()

# =========================================================
# ТЕСТОВАЯ ВЫДАЧА ДОСТУПА К ПЛАТНОЙ ТЕМЕ
# =========================================================

@router.post("/users/{user_id}/topics/{topic_id}/grant-access")
def grant_topic_access(
    user_id: int,
    topic_id: int,
    data: GrantTopicAccess,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        if data.minutes < 1:
            raise HTTPException(
                status_code=400,
                detail="Количество минут должно быть не меньше 1"
            )

        user = db.query(User).filter(
            User.id == user_id
        ).first()

        if not user:
            raise HTTPException(
                status_code=404,
                detail="Пользователь не найден"
            )

        topic = db.query(Topic).filter(
            Topic.id == topic_id
        ).first()

        if not topic:
            raise HTTPException(
                status_code=404,
                detail="Тема не найдена"
            )

        if user.service_id != topic.service_id:
            raise HTTPException(
                status_code=400,
                detail="Тема не относится к службе пользователя"
            )

        if (topic.price_kopecks or 0) <= 0:
            raise HTTPException(
                status_code=400,
                detail="Эта тема бесплатная и не требует временного доступа"
            )

        now = datetime.utcnow()

        current_access = (
            db.query(TopicAccess)
            .filter(
                TopicAccess.user_id == user.id,
                TopicAccess.topic_id == topic.id,
                TopicAccess.expires_at > now
            )
            .order_by(TopicAccess.expires_at.desc())
            .first()
        )

        # Если доступ уже действует, добавляем минуты к текущему окончанию.
        # Если доступа нет/он истёк — считаем от текущего момента.
        starts_from = (
            current_access.expires_at
            if current_access
            else now
        )

        expires_at = starts_from + timedelta(
            minutes=data.minutes
        )

        access = TopicAccess(
            user_id=user.id,
            topic_id=topic.id,
            granted_at=now,
            expires_at=expires_at,
            source="admin_test"
        )

        db.add(access)
        db.commit()
        db.refresh(access)

        return {
            "status": "granted",
            "access_id": access.id,
            "user_id": user.id,
            "topic_id": topic.id,
            "topic_name": topic.name,
            "minutes": data.minutes,
            "granted_at": access.granted_at.isoformat(),
            "expires_at": access.expires_at.isoformat()
        }

    finally:
        db.close()


# =========================================================
# ВОПРОСЫ
# =========================================================

class QuestionCreate(BaseModel):
    service_id: int
    topic_id: int
    question: str
    answer: str
    keywords: str | None = None

class QuestionUpdate(BaseModel):
    service_id: int
    topic_id: int
    question: str
    answer: str
    keywords: str | None = None

@router.get("/questions")
def get_questions(
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        questions = (
            db.query(Question)
            .order_by(Question.id.desc())
            .all()
        )

        result = []

        for question in questions:

            service = db.query(Service).filter(
                Service.id == question.service_id
            ).first()

            topic = db.query(Topic).filter(
                Topic.id == question.topic_id
            ).first()

            result.append({
                "id": question.id,
                "service_id": question.service_id,
                "service_name": (
                    service.name if service else ""
                ),
                "topic_id": question.topic_id,
                "topic_name": (
                    topic.name if topic else ""
                ),
                "question": question.question,
                "answer": question.answer,
                "keywords": question.keywords
            })

        return result

    finally:
        db.close()


@router.post("/questions")
def create_question(
    data: QuestionCreate,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        service = db.query(Service).filter(
            Service.id == data.service_id
        ).first()

        if not service:
            raise HTTPException(
                status_code=404,
                detail="Служба не найдена"
            )

        topic = db.query(Topic).filter(
            Topic.id == data.topic_id
        ).first()

        if not topic:
            raise HTTPException(
                status_code=404,
                detail="Тема не найдена"
            )

        if topic.service_id != service.id:
            raise HTTPException(
                status_code=400,
                detail="Эта тема не относится к выбранной службе"
            )

        question_text = data.question.strip()
        answer_text = data.answer.strip()

        if not question_text:
            raise HTTPException(
                status_code=400,
                detail="Введите вопрос"
            )

        if not answer_text:
            raise HTTPException(
                status_code=400,
                detail="Введите правильный ответ"
            )

        question = Question(
            service_id=service.id,
            topic_id=topic.id,
            question=question_text,
            answer=answer_text,
            keywords=(
                data.keywords.strip()
                if data.keywords
                else None
            )
        )

        db.add(question)
        db.commit()
        db.refresh(question)

        return {
            "id": question.id,
            "service_id": question.service_id,
            "topic_id": question.topic_id,
            "question": question.question,
            "answer": question.answer,
            "keywords": question.keywords
        }

    finally:
        db.close()


@router.delete("/questions/{question_id}")
def delete_question(
    question_id: int,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        question = db.query(Question).filter(
            Question.id == question_id
        ).first()

        if not question:
            raise HTTPException(
                status_code=404,
                detail="Вопрос не найден"
            )

        db.delete(question)
        db.commit()

        return {
            "status": "deleted"
        }

    finally:
        db.close()

@router.put("/questions/{question_id}")
def update_question(
    question_id: int,
    data: QuestionUpdate,
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        question = db.query(Question).filter(
            Question.id == question_id
        ).first()

        if not question:
            raise HTTPException(
                status_code=404,
                detail="Вопрос не найден"
            )

        service = db.query(Service).filter(
            Service.id == data.service_id
        ).first()

        if not service:
            raise HTTPException(
                status_code=404,
                detail="Служба не найдена"
            )

        topic = db.query(Topic).filter(
            Topic.id == data.topic_id
        ).first()

        if not topic:
            raise HTTPException(
                status_code=404,
                detail="Тема не найдена"
            )

        if topic.service_id != service.id:
            raise HTTPException(
                status_code=400,
                detail="Эта тема не относится к выбранной службе"
            )

        question_text = data.question.strip()
        answer_text = data.answer.strip()

        if not question_text:
            raise HTTPException(
                status_code=400,
                detail="Введите вопрос"
            )

        if not answer_text:
            raise HTTPException(
                status_code=400,
                detail="Введите правильный ответ"
            )

        question.service_id = service.id
        question.topic_id = topic.id
        question.question = question_text
        question.answer = answer_text
        question.keywords = (
            data.keywords.strip()
            if data.keywords
            else None
        )

        db.commit()
        db.refresh(question)

        return {
            "status": "updated",
            "id": question.id,
            "service_id": question.service_id,
            "topic_id": question.topic_id,
            "question": question.question,
            "answer": question.answer,
            "keywords": question.keywords
        }

    finally:
        db.close()


@router.post("/questions/import")
async def import_questions(
    file: UploadFile = File(...),
    admin=Depends(require_admin)
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Поддерживаются только файлы .xlsx"
        )

    content = await file.read()

    try:
        workbook = load_workbook(
            filename=BytesIO(content),
            data_only=True
        )
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Не удалось прочитать Excel-файл"
        )

    sheet = workbook.active

    db = SessionLocal()

    imported = 0
    skipped = []

    try:
        for row_number, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            service_name, topic_name, question_text, answer_text, keywords = (
                list(row) + [None] * 5
            )[:5]

            if not any(row):
                continue

            if (
                not service_name
                or not topic_name
                or not question_text
                or not answer_text
            ):
                skipped.append({
                    "row": row_number,
                    "reason": "Не заполнены обязательные поля"
                })
                continue

            service = db.query(Service).filter(
                Service.name == str(service_name).strip()
            ).first()

            if not service:
                skipped.append({
                    "row": row_number,
                    "reason": f"Служба не найдена: {service_name}"
                })
                continue

            topic = db.query(Topic).filter(
                Topic.name == str(topic_name).strip(),
                Topic.service_id == service.id
            ).first()

            if not topic:
                skipped.append({
                    "row": row_number,
                    "reason": f"Тема не найдена: {topic_name}"
                })
                continue

            exists = db.query(Question).filter(
                Question.service_id == service.id,
                Question.topic_id == topic.id,
                Question.question == str(question_text).strip()
            ).first()

            if exists:
                skipped.append({
                    "row": row_number,
                    "reason": "Такой вопрос уже существует"
                })
                continue

            question = Question(
                service_id=service.id,
                topic_id=topic.id,
                question=str(question_text).strip(),
                answer=str(answer_text).strip(),
                keywords=(
                    str(keywords).strip()
                    if keywords
                    else None
                )
            )

            db.add(question)
            imported += 1

        db.commit()

        return {
            "imported": imported,
            "skipped_count": len(skipped),
            "skipped": skipped
        }

    finally:
        db.close()

@router.get("/questions/export")
def export_questions(
    admin=Depends(require_admin)
):
    db = SessionLocal()

    try:
        questions = (
            db.query(Question, Service, Topic)
            .join(
                Service,
                Question.service_id == Service.id
            )
            .join(
                Topic,
                Question.topic_id == Topic.id
            )
            .order_by(
                Service.name,
                Topic.name,
                Question.id
            )
            .all()
        )

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Вопросы"

        sheet.append([
            "Служба",
            "Тема",
            "Вопрос",
            "Правильный ответ",
            "Ключевые слова"
        ])

        for question, service, topic in questions:
            sheet.append([
                service.name,
                topic.name,
                question.question,
                question.answer,
                question.keywords or ""
            ])

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        headers = {
            "Content-Disposition":
                'attachment; filename="TestMaster_questions.xlsx"'
        }

        return StreamingResponse(
            output,
            media_type=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers=headers
        )

    finally:
        db.close()    